from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


METRIC_ALIASES = {
    'target_achievement': ('target_achievement', 'target achievement'),
    'productivity': ('productivity',),
    'quality': ('quality',),
    'punctuality': ('punctuality', 'late login', 'late login / punctuality'),
    'attendance': ('attendance', 'absent', 'absence'),
    'competency': ('competency', 'competencies', 'behavioral competency', 'behavioural competency'),
}

TITLE_FILL = '17365D'
HEADER_FILL = 'DCE6F1'
TENURE_FILL = 'F2DCDB'
TOTAL_FILL = 'E2F0C2'
WHITE = 'FFFFFF'
BLACK = '000000'


def fiscal_year_label(year: int, month: int, *, compact: bool = False) -> str:
    start_year = year if month >= 4 else year - 1
    end_year = start_year + 1
    if compact:
        return f'{str(start_year)[-2:]}-{str(end_year)[-2:]}'
    return f'{start_year}-{str(end_year)[-2:]}'


def metric_for(row, key: str):
    aliases = METRIC_ALIASES[key]
    for metric in row.metrics:
        metric_key = str(metric.metric_key or '').strip().lower()
        metric_name = str(metric.metric_name_snapshot or '').strip().lower()
        if metric_key in aliases or metric_name in aliases:
            return metric
    for metric in row.metrics:
        metric_key = str(metric.metric_key or '').strip().lower()
        metric_name = str(metric.metric_name_snapshot or '').strip().lower()
        if any(alias in metric_key or alias in metric_name for alias in aliases):
            return metric
    return None


def metric_value(row, key: str, target_achievement_percent: float | None = None):
    metric = metric_for(row, key)
    if not metric:
        return None
    if key == 'target_achievement' and target_achievement_percent is not None:
        weight = float(metric.weight_snapshot or 0)
        percent = max(0.0, min(float(target_achievement_percent), 100.0))
        return round(max(0.0, min((weight * percent) / 100, weight)), 2)
    return round(float(metric.final_value or 0), 2)


def month_token(year: int, month: int) -> str:
    return f"{['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'][month - 1]}'{str(year)[-2:]}"


def row_value(row, name: str):
    return getattr(row, name, None) or ''


def export_monthly_tables(tables, *, target_achievement_percent_by_period: dict[tuple[int, int], float] | None = None) -> BytesIO:
    tables = list(tables)
    if not tables:
        raise ValueError('At least one PMS table is required')

    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    sheet = workbook.active
    sheet.title = 'PMS Monthly Data'
    sheet.sheet_view.showGridLines = False

    first_table = tables[0]
    last_table = tables[-1]
    if first_table.year == last_table.year and first_table.month == last_table.month:
        period_label = fiscal_year_label(first_table.year, first_table.month, compact=True)
    else:
        period_label = f'{month_token(first_table.year, first_table.month)} to {month_token(last_table.year, last_table.month)}'
    title = f'PMS Monthly Data {period_label} For L-1 & L-2'
    sheet.merge_cells('A1:P1')
    sheet['A1'] = title

    sheet.merge_cells('A2:A4')
    sheet.merge_cells('B2:H2')
    sheet.merge_cells('I2:K2')
    sheet.merge_cells('L2:N2')
    sheet.merge_cells('O2:O3')
    sheet.merge_cells('P2:P4')
    for cell_range in ['B3:B4', 'C3:C4', 'D3:D4', 'E3:E4', 'F3:F4', 'G3:G4', 'H3:H4', 'I3:I4', 'J3:J4', 'K3:K4', 'N3:N4']:
        sheet.merge_cells(cell_range)
    sheet.merge_cells('L3:M3')

    header_values = {
        'A2': 'Sr No',
        'B2': 'Employee Details',
        'B3': 'Employee\nCode',
        'C3': 'Employee Name',
        'D3': 'Department',
        'E3': 'Designation',
        'F3': 'DOJ',
        'G3': 'Tenure',
        'H3': 'Month',
        'I2': 'Performance Outcomes',
        'I3': 'Target\nAchievement\n65%',
        'J3': 'Productivity\n10%',
        'K3': 'Quality 10%',
        'L2': 'Behavioral Competencies',
        'L3': 'Attendance & Punctuality 10%',
        'L4': 'Late Login 5%',
        'M4': 'Absent 5%',
        'N3': 'Competency\n5%',
        'O2': 'Total Weightage',
        'O4': '100%',
        'P2': 'Remarks',
    }
    for cell, value in header_values.items():
        sheet[cell] = value

    widths = {
        'A': 8,
        'B': 13,
        'C': 30,
        'D': 18,
        'E': 34,
        'F': 12,
        'G': 10,
        'H': 10,
        'I': 14,
        'J': 13,
        'K': 13,
        'L': 14,
        'M': 14,
        'N': 13,
        'O': 17,
        'P': 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.row_dimensions[1].height = 29
    sheet.row_dimensions[2].height = 25
    sheet.row_dimensions[3].height = 31
    sheet.row_dimensions[4].height = 27

    thin = Side(style='thin', color=BLACK)
    medium = Side(style='medium', color=BLACK)
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)

    title_cell = sheet['A1']
    title_cell.fill = PatternFill('solid', fgColor=TITLE_FILL)
    title_cell.font = Font(name='Times New Roman', size=14, bold=True, color=WHITE)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=16):
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=HEADER_FILL)
            cell.font = Font(name='Times New Roman', size=10, bold=True, color=BLACK)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    for cell in sheet[1]:
        cell.fill = PatternFill('solid', fgColor=TITLE_FILL)
        cell.border = medium_border

    row_index = 5
    serial_number = 1
    target_achievement_percent_by_period = target_achievement_percent_by_period or {}
    for table in tables:
        target_achievement_percent = target_achievement_percent_by_period.get((table.year, table.month))
        for row in table.items:
            values = {
                'A': serial_number,
                'B': row_value(row, 'employee_id'),
                'C': row.user_name or '',
                'D': row_value(row, 'department'),
                'E': row_value(row, 'designation'),
                'F': getattr(row, 'date_of_joining', None),
                'G': '',
                'H': month_token(table.year, table.month),
                'I': metric_value(row, 'target_achievement', target_achievement_percent),
                'J': metric_value(row, 'productivity'),
                'K': metric_value(row, 'quality'),
                'L': metric_value(row, 'punctuality'),
                'M': metric_value(row, 'attendance'),
                'N': metric_value(row, 'competency'),
                'O': f'=SUM(I{row_index}:N{row_index})',
                'P': getattr(row, 'remarks', None) or '',
            }
            for column, value in values.items():
                cell = sheet[f'{column}{row_index}']
                cell.value = '' if value is None else value
                cell.font = Font(name='Times New Roman', size=10, color=BLACK)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if column == 'F' and value:
                    cell.number_format = 'm/d/yyyy'
            sheet[f'C{row_index}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet[f'P{row_index}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet[f'G{row_index}'].fill = PatternFill('solid', fgColor=TENURE_FILL)
            sheet[f'O{row_index}'].fill = PatternFill('solid', fgColor=TOTAL_FILL)
            sheet.row_dimensions[row_index].height = 20
            row_index += 1
            serial_number += 1

    last_row = max(row_index - 1, 5)
    for row in sheet.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=16):
        for cell in row:
            if cell.row == 1:
                continue
            cell.border = thin_border
    for row in range(1, last_row + 1):
        for column in (1, 8, 11, 14, 16):
            sheet.cell(row=row, column=column).border = Border(
                left=sheet.cell(row=row, column=column).border.left,
                right=medium,
                top=sheet.cell(row=row, column=column).border.top,
                bottom=sheet.cell(row=row, column=column).border.bottom,
            )

    sheet.freeze_panes = 'A5'
    sheet.print_title_rows = '1:4'
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_monthly_table(table, *, target_achievement_percent: float | None = None) -> BytesIO:
    percent_by_period = (
        {(table.year, table.month): target_achievement_percent}
        if target_achievement_percent is not None
        else None
    )
    return export_monthly_tables([table], target_achievement_percent_by_period=percent_by_period)
