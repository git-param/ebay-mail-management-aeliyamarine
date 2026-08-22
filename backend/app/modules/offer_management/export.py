from io import BytesIO
from datetime import date, datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_COLUMNS = [
    ('Entry Number', 'entry_number'),
    ('Offer Date', 'offer_date'),
    ('Agent', 'agent_name'),
    ('eBay Account', 'ebay_account_name'),
    ('Listing ID', 'listing_id'),
    ('SKU', 'sku'),
    ('Product Title', 'product_title'),
    ('Condition', 'condition'),
    ('Listing Quantity', 'listing_quantity'),
    ('Offer Quantity', 'offer_quantity'),
    ('Currency', 'currency'),
    ('Listed Price', 'listed_price'),
    ('Revised Price', 'revised_price'),
    ('Automated Offer', 'automated_offer_price'),
    ('Buyer Offer', 'buyer_offer_price'),
    ('Counteroffer/Best Price', 'counteroffer_price'),
    ('Final Price', 'final_price'),
    ('Buyer ID', 'buyer_id'),
    ('Status', 'status'),
    ('Outcome', 'outcome'),
    ('High Value', 'is_high_value'),
    ('VIP Lead', 'is_vip_lead'),
    ('Next Offer Follow-up', 'next_offer_followup'),
    ('Follow-up 1', 'follow_up_1_notes'),
    ('Follow-up 2', 'follow_up_2_notes'),
    ('Remarks', 'remarks'),
    ('eBay Link', 'listing_url'),
    ('Created At', 'created_at'),
    ('Updated At', 'updated_at'),
]


def normalize_excel_value(value):
    if hasattr(value, 'value'):
        value = value.value
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def export_entries(entries) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Offer Entries'
    sheet.append([label for label, _ in EXPORT_COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F2937')
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:{get_column_letter(len(EXPORT_COLUMNS))}1'
    money_fields = {'listed_price', 'revised_price', 'automated_offer_price', 'buyer_offer_price', 'counteroffer_price', 'final_price'}
    wrap_fields = {'product_title', 'remarks', 'follow_up_1_notes', 'follow_up_2_notes'}
    for entry in entries:
        row = []
        for _, field in EXPORT_COLUMNS:
            value = getattr(entry, field, None)
            if field == 'agent_name':
                value = entry.created_by.full_name if entry.created_by else ''
            row.append(normalize_excel_value(value))
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row, start=1):
            field = EXPORT_COLUMNS[index - 1][1]
            if field in money_fields and cell.value is not None:
                cell.number_format = '#,##0.00'
            if field in wrap_fields:
                cell.alignment = cell.alignment.copy(wrap_text=True)
    widths = {
        'A': 14, 'B': 14, 'C': 22, 'D': 22, 'E': 16, 'F': 18, 'G': 42,
        'H': 18, 'I': 14, 'J': 14, 'K': 10, 'L': 14, 'M': 14, 'N': 16,
        'O': 14, 'P': 22, 'Q': 14, 'R': 18, 'S': 24, 'T': 24, 'U': 12,
        'V': 10, 'W': 16, 'X': 18, 'Y': 30, 'Z': 16, 'AA': 18, 'AB': 30,
        'AC': 36, 'AD': 34, 'AE': 22, 'AF': 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
