"""Enterprise analytics and reporting services for conversation operations."""

from __future__ import annotations

import tempfile
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, exists, func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.api.dependencies import is_support_agent
from app.models.category import Category
from app.models.conversation import (
    Conversation,
    ConversationAssignment,
    ConversationStatus,
    ConversationStatusHistory,
    Message,
    MessageSenderType,
)


@dataclass(slots=True)
class AnalyticsFilters:
    """Validated filter values shared by dashboards and exports."""

    start_date: date | None = None
    end_date: date | None = None
    agent_id: UUID | None = None
    category_id: UUID | None = None
    status: ConversationStatus | None = None


class AnalyticsService:
    """Calculate operational analytics from normalized conversation history."""

    def __init__(self, db: Session):
        """
        Initialize the analytics service.

        Purpose:
        Provides a single source of truth for dashboard metrics and Excel
        export calculations.

        Parameters:
        db: Active SQLAlchemy session used for read-only reporting queries.

        Returns:
        A service instance bound to the current request database session.

        Business Logic:
        Reporting is calculated from conversations, messages, assignments,
        and status history so historical reporting remains reproducible.
        """
        self.db = db

    def dashboard(self, filters: AnalyticsFilters, current_user) -> dict:
        """
        Build role-aware dashboard analytics for the requested filter set.

        Purpose:
        Aggregates queue, productivity, category, SLA, and response-time
        metrics for admins, operations managers, and support agents.

        Parameters:
        filters: Date range, agent, category, and status filters.
        current_user: Authenticated user used to enforce agent scoping.

        Returns:
        A serializable dictionary matching AnalyticsDashboardResponse.

        Business Logic:
        Support agents are always scoped to their own currently assigned
        conversations and replies. Admin and operations users may filter by
        any agent.

        Important Notes:
        SLA compliance is based on the first agent response after the first
        inbound buyer message, falling back to open SLA deadline breach checks
        when no response exists.
        """
        scoped_filters = self._scope_filters(filters, current_user)
        conversations = self._conversation_rows(scoped_filters)
        rows = [self._conversation_record(conversation) for conversation in conversations]
        total_replies = sum(row['reply_count'] for row in rows)
        pending = [row for row in rows if row['status'] == ConversationStatus.PENDING.value]
        overdue = [row for row in rows if row['is_overdue']]
        compliant_count = sum(1 for row in rows if row['sla_compliant'] is True)
        measured_sla = sum(1 for row in rows if row['sla_compliant'] is not None)
        avg_response = self._average([row['first_response_minutes'] for row in rows if row['first_response_minutes'] is not None])

        return {
            'role_scope': self._role_scope(current_user),
            'totals': [
                self._metric('Total conversations', len(rows)),
                self._metric('Total replies', total_replies),
                self._metric('Pending conversations', len(pending)),
                self._metric('Overdue conversations', len(overdue)),
                self._metric('SLA compliance', self._percent(compliant_count, measured_sla)),
                self._metric('Average response time', self._minutes_label(avg_response)),
            ],
            'by_category': self._count_by(rows, 'category_name', default='Uncategorized'),
            'by_status': self._count_by(rows, 'status', default='Unknown'),
            'by_assigned_user': self._count_by(rows, 'assigned_to_name', default='Unassigned'),
            'daily_trends': self._daily_trends(rows, scoped_filters),
            'sla_metrics': [
                self._metric('Compliant', compliant_count),
                self._metric('Breached', max(measured_sla - compliant_count, 0)),
                self._metric('Open overdue', len(overdue)),
            ],
            'agent_productivity': self._count_by(rows, 'assigned_to_name', default='Unassigned'),
            'category_distribution': self._count_by(rows, 'category_name', default='Uncategorized'),
            'agent_summary': self._agent_summary(rows),
            'category_summary': self._category_summary(rows),
            'filters': {
                'start_date': scoped_filters.start_date.isoformat() if scoped_filters.start_date else None,
                'end_date': scoped_filters.end_date.isoformat() if scoped_filters.end_date else None,
                'agent_id': str(scoped_filters.agent_id) if scoped_filters.agent_id else None,
                'category_id': str(scoped_filters.category_id) if scoped_filters.category_id else None,
                'status': scoped_filters.status.value if scoped_filters.status else None,
            },
        }

    def export_workbook(self, filters: AnalyticsFilters, current_user) -> Path:
        """
        Generate an Excel workbook with raw data, summaries, SLA, and charts.

        Purpose:
        Creates a business-ready XLSX report for operational review and
        offline analysis.

        Parameters:
        filters: Dashboard filters to apply to the exported data.
        current_user: Authenticated user used for role-based reporting scope.

        Returns:
        Path to a temporary XLSX file ready to stream with FileResponse.

        Business Logic:
        Uses the same dashboard calculation path to avoid duplicate totals.
        Charts are embedded in the workbook's Charts sheet.

        Important Notes:
        The caller is responsible for returning or deleting the generated
        temporary file after the HTTP response lifecycle completes.
        """
        scoped_filters = self._scope_filters(filters, current_user)
        rows = [self._conversation_record(conversation) for conversation in self._conversation_rows(scoped_filters)]
        workbook = Workbook()
        raw_sheet = workbook.active
        raw_sheet.title = 'Raw Data'
        self._write_raw_data(raw_sheet, rows)
        agent_sheet = workbook.create_sheet('Agent Summary')
        self._write_summary(agent_sheet, self._agent_summary(rows), ['Agent', 'Conversations', 'Replies', 'Pending', 'Overdue', 'Average Response'])
        category_sheet = workbook.create_sheet('Category Summary')
        self._write_summary(category_sheet, self._category_summary(rows), ['Category', 'Conversations', 'Replies', 'Pending', 'Overdue', 'SLA Compliance'])
        sla_sheet = workbook.create_sheet('SLA Summary')
        self._write_summary(sla_sheet, self._sla_summary(rows), ['Metric', 'Value'])
        charts_sheet = workbook.create_sheet('Charts')
        self._write_charts(charts_sheet, agent_sheet, category_sheet, sla_sheet, rows)
        path = Path(tempfile.NamedTemporaryFile(prefix='conversation_report_', suffix='.xlsx', delete=False).name)
        workbook.save(path)
        return path

    def _conversation_rows(self, filters: AnalyticsFilters) -> list[Conversation]:
        """
        Return conversations matching report filters with related history loaded.

        Purpose:
        Centralizes the reporting query so dashboards and exports read the
        same filtered conversation population.

        Parameters:
        filters: Normalized date, agent, category, and status filters.

        Returns:
        List of Conversation ORM objects with messages, assignments, status
        history, and categories eagerly loaded.

        Business Logic:
        Date filters apply to the latest message timestamp when available,
        otherwise conversation creation time. Agent filters match current
        assignments or replies authored by that agent.
        """
        statement = (
            select(Conversation)
            .options(
                selectinload(Conversation.messages),
                selectinload(Conversation.assignments).selectinload(ConversationAssignment.assignee),
                selectinload(Conversation.assignments).selectinload(ConversationAssignment.assigner),
                selectinload(Conversation.status_history).selectinload(ConversationStatusHistory.user),
                selectinload(Conversation.category),
            )
            .order_by(Conversation.last_message_at.desc().nullslast(), Conversation.created_at.desc())
        )
        report_timestamp = func.coalesce(Conversation.last_message_at, Conversation.created_at)
        if filters.start_date:
            statement = statement.where(report_timestamp >= self._start_datetime(filters.start_date))
        if filters.end_date:
            statement = statement.where(report_timestamp <= self._end_datetime(filters.end_date))
        if filters.category_id:
            statement = statement.where(Conversation.category_id == filters.category_id)
        if filters.status:
            statement = statement.where(Conversation.status == filters.status)
        if filters.agent_id:
            assignment_filter = exists(
                select(ConversationAssignment.id).where(
                    and_(
                        ConversationAssignment.conversation_id == Conversation.id,
                        ConversationAssignment.assigned_to == filters.agent_id,
                        ConversationAssignment.unassigned_at.is_(None),
                    )
                )
            )
            reply_filter = exists(
                select(Message.id).where(
                    and_(
                        Message.conversation_id == Conversation.id,
                        Message.sender_type == MessageSenderType.AGENT,
                        Message.raw_payload['actor_id'].as_string() == str(filters.agent_id),
                    )
                )
            )
            statement = statement.where(or_(assignment_filter, reply_filter))
        return list(self.db.scalars(statement))

    def _conversation_record(self, conversation: Conversation) -> dict:
        """
        Convert one conversation into a report-ready historical record.

        Purpose:
        Derives audit and SLA fields without mutating the conversation.

        Parameters:
        conversation: Conversation ORM object with related collections loaded.

        Returns:
        Dictionary containing raw report columns and intermediate metrics.

        Business Logic:
        The latest message determines direction; unread inbound buyer
        messages become Not Read; any agent reply after buyer contact marks
        the conversation as Replied.
        """
        messages = sorted(conversation.messages, key=lambda item: item.sent_at)
        assignments = sorted(conversation.assignments, key=lambda item: item.assigned_at)
        latest = messages[-1] if messages else None
        inbound_messages = [message for message in messages if message.is_inbound or message.sender_type == MessageSenderType.CUSTOMER]
        reply_messages = [message for message in messages if message.sender_type == MessageSenderType.AGENT or not message.is_inbound]
        first_inbound_at = inbound_messages[0].sent_at if inbound_messages else None
        first_response = self._first_response_after(reply_messages, first_inbound_at)
        current_assignment = next((assignment for assignment in reversed(assignments) if assignment.unassigned_at is None), None)
        closed_history = next(
            (history for history in sorted(conversation.status_history, key=lambda item: item.changed_at, reverse=True) if history.new_status in {ConversationStatus.CLOSED, ConversationStatus.RESOLVED}),
            None,
        )
        resolution_at = closed_history.changed_at if closed_history else None
        sla_hours = conversation.category.sla_hours if conversation.category else 24
        due_at = first_inbound_at + timedelta(hours=sla_hours) if first_inbound_at else None
        is_replied = bool(reply_messages)
        is_not_read = bool(latest and latest.is_inbound and (latest.read_status is False or conversation.unread_count > 0))
        calculated_status = 'Not Read' if is_not_read else 'Replied' if is_replied else conversation.status.value
        first_response_minutes = (
            (first_response.sent_at - first_inbound_at).total_seconds() / 60
            if first_response and first_inbound_at
            else None
        )
        sla_compliant = None
        if first_inbound_at:
            if first_response:
                sla_compliant = bool(due_at and first_response.sent_at <= due_at)
            elif conversation.status not in {ConversationStatus.CLOSED, ConversationStatus.RESOLVED}:
                sla_compliant = bool(due_at and datetime.now(UTC) <= due_at)
        return {
            'conversation_id': str(conversation.id),
            'provider_conversation_id': conversation.provider_conversation_id,
            'subject': conversation.subject,
            'buyer': conversation.buyer_identifier,
            'category_name': conversation.category.name if conversation.category else None,
            'status': conversation.status.value,
            'calculated_status': calculated_status,
            'last_message_direction': self._message_direction(latest),
            'message_count': len(messages),
            'reply_count': len(reply_messages),
            'assigned_by_name': current_assignment.assigner.full_name if current_assignment and current_assignment.assigner else None,
            'assigned_to_name': current_assignment.assignee.full_name if current_assignment and current_assignment.assignee else None,
            'assigned_at': current_assignment.assigned_at if current_assignment else None,
            'replied_by_name': self._reply_actor_name(reply_messages[-1]) if reply_messages else None,
            'reply_timestamp': reply_messages[-1].sent_at if reply_messages else None,
            'first_response_timestamp': first_response.sent_at if first_response else None,
            'first_response_minutes': first_response_minutes,
            'closed_by_name': closed_history.user.full_name if closed_history and closed_history.user else None,
            'resolution_timestamp': resolution_at,
            'last_message_at': conversation.last_message_at,
            'created_at': conversation.created_at,
            'sla_due_at': due_at,
            'sla_compliant': sla_compliant,
            'is_replied': is_replied,
            'is_not_read': is_not_read,
            'is_overdue': bool(due_at and not first_response and conversation.status not in {ConversationStatus.CLOSED, ConversationStatus.RESOLVED} and due_at < datetime.now(UTC)),
        }

    def _write_raw_data(self, sheet, rows: list[dict]) -> None:
        """
        Write row-level conversation facts to the Raw Data worksheet.

        Purpose:
        Gives reporting users an auditable basis for every summary metric.

        Parameters:
        sheet: openpyxl worksheet to populate.
        rows: Report records produced by _conversation_record.

        Returns:
        None.

        Business Logic:
        Includes assignment, reply, first-response, and resolution timestamps
        required for complete historical reporting.
        """
        headers = [
            'Conversation ID', 'Provider Conversation ID', 'Subject', 'Buyer', 'Category', 'Status',
            'Calculated Status', 'Last Message Direction', 'Messages', 'Replies', 'Assigned By',
            'Assigned To', 'Assigned At', 'Replied By', 'Reply Timestamp', 'First Response Timestamp',
            'Resolution Timestamp', 'Closed By', 'SLA Due At', 'SLA Compliant', 'Overdue',
        ]
        sheet.append(headers)
        for row in rows:
            sheet.append([
                row['conversation_id'], row['provider_conversation_id'], row['subject'], row['buyer'],
                row['category_name'] or 'Uncategorized', row['status'], row['calculated_status'],
                row['last_message_direction'], row['message_count'], row['reply_count'],
                row['assigned_by_name'], row['assigned_to_name'], self._excel_datetime(row['assigned_at']),
                row['replied_by_name'], self._excel_datetime(row['reply_timestamp']), self._excel_datetime(row['first_response_timestamp']),
                self._excel_datetime(row['resolution_timestamp']), row['closed_by_name'], self._excel_datetime(row['sla_due_at']),
                'Yes' if row['sla_compliant'] else 'No' if row['sla_compliant'] is False else 'N/A',
                'Yes' if row['is_overdue'] else 'No',
            ])
        self._style_header(sheet)

    def _write_summary(self, sheet, rows: list[dict], headers: list[str]) -> None:
        """
        Write a generic summary table to a worksheet.

        Purpose:
        Reuses consistent header styling and row appending across summary
        sheets.

        Parameters:
        sheet: openpyxl worksheet to populate.
        rows: List of dictionaries keyed by the supplied headers.
        headers: Ordered header names to write.

        Returns:
        None.

        Business Logic:
        Keeps exported summary columns stable for downstream reporting users.
        """
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        self._style_header(sheet)

    def _write_charts(self, sheet, agent_sheet, category_sheet, sla_sheet, rows: list[dict]) -> None:
        """
        Build embedded workbook charts from summary worksheet tables.

        Purpose:
        Adds Excel-native charts for productivity, category distribution, SLA,
        and daily trend review.

        Parameters:
        sheet: Destination Charts worksheet.
        agent_sheet: Agent Summary worksheet.
        category_sheet: Category Summary worksheet.
        sla_sheet: SLA Summary worksheet.
        rows: Report records used to write daily trend chart source data on
        the Charts sheet.

        Returns:
        None.

        Business Logic:
        Charts bind to generated summary ranges. Daily trend source data is
        written directly on the Charts sheet to preserve the requested workbook
        sheet list.
        """
        sheet['A1'] = 'Charts'
        sheet['A1'].font = Font(bold=True, size=16)
        if agent_sheet.max_row > 1:
            chart = BarChart()
            chart.title = 'Agent Productivity'
            chart.add_data(Reference(agent_sheet, min_col=2, min_row=1, max_row=agent_sheet.max_row), titles_from_data=True)
            chart.set_categories(Reference(agent_sheet, min_col=1, min_row=2, max_row=agent_sheet.max_row))
            sheet.add_chart(chart, 'A3')
        if category_sheet.max_row > 1:
            chart = BarChart()
            chart.title = 'Category Distribution'
            chart.add_data(Reference(category_sheet, min_col=2, min_row=1, max_row=category_sheet.max_row), titles_from_data=True)
            chart.set_categories(Reference(category_sheet, min_col=1, min_row=2, max_row=category_sheet.max_row))
            sheet.add_chart(chart, 'J3')
        if sla_sheet.max_row > 1:
            chart = PieChart()
            chart.title = 'SLA Compliance'
            chart.add_data(Reference(sla_sheet, min_col=2, min_row=2, max_row=min(sla_sheet.max_row, 3)))
            chart.set_categories(Reference(sla_sheet, min_col=1, min_row=2, max_row=min(sla_sheet.max_row, 3)))
            sheet.add_chart(chart, 'A20')
        daily_trends = self._daily_trends(rows, AnalyticsFilters())
        if daily_trends:
            sheet['J18'] = 'Date'
            sheet['K18'] = 'Conversations'
            for index, item in enumerate(daily_trends, start=19):
                sheet.cell(row=index, column=10, value=item['label'])
                sheet.cell(row=index, column=11, value=item['value'])
            chart = LineChart()
            chart.title = 'Daily Trend Analysis'
            chart.add_data(Reference(sheet, min_col=11, min_row=18, max_row=18 + len(daily_trends)), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=10, min_row=19, max_row=18 + len(daily_trends)))
            sheet.add_chart(chart, 'J20')

    def _agent_summary(self, rows: list[dict]) -> list[dict]:
        """
        Calculate agent-level handling metrics.

        Purpose:
        Summarizes conversation volume, reply volume, pending workload, overdue
        workload, and response speed by current assignee.

        Parameters:
        rows: Report records produced by _conversation_record.

        Returns:
        List of dictionaries ready for API serialization or Excel export.

        Business Logic:
        Unassigned conversations are retained under an Unassigned bucket so
        operational workload does not disappear from reports.
        """
        grouped = self._group_rows(rows, 'assigned_to_name', 'Unassigned')
        return [
            {
                'Agent': label,
                'Conversations': len(items),
                'Replies': sum(item['reply_count'] for item in items),
                'Pending': sum(1 for item in items if item['status'] == ConversationStatus.PENDING.value),
                'Overdue': sum(1 for item in items if item['is_overdue']),
                'Average Response': self._minutes_label(self._average([item['first_response_minutes'] for item in items if item['first_response_minutes'] is not None])),
            }
            for label, items in grouped.items()
        ]

    def _category_summary(self, rows: list[dict]) -> list[dict]:
        """
        Calculate category-level handling metrics.

        Purpose:
        Shows category-wise conversation load, replies, pending items, overdue
        items, and SLA compliance.

        Parameters:
        rows: Report records produced by _conversation_record.

        Returns:
        List of dictionaries ready for API serialization or Excel export.

        Business Logic:
        Uncategorized conversations are retained for queue hygiene analysis.
        """
        grouped = self._group_rows(rows, 'category_name', 'Uncategorized')
        return [
            {
                'Category': label,
                'Conversations': len(items),
                'Replies': sum(item['reply_count'] for item in items),
                'Pending': sum(1 for item in items if item['status'] == ConversationStatus.PENDING.value),
                'Overdue': sum(1 for item in items if item['is_overdue']),
                'SLA Compliance': self._percent(
                    sum(1 for item in items if item['sla_compliant'] is True),
                    sum(1 for item in items if item['sla_compliant'] is not None),
                ),
            }
            for label, items in grouped.items()
        ]

    def _sla_summary(self, rows: list[dict]) -> list[dict]:
        """
        Calculate SLA summary rows for Excel output.

        Purpose:
        Provides compact SLA counts for charting and executive review.

        Parameters:
        rows: Report records produced by _conversation_record.

        Returns:
        List of metric dictionaries with Metric and Value keys.

        Business Logic:
        Only conversations with a measurable inbound buyer message are counted
        in compliant and breached SLA buckets.
        """
        compliant = sum(1 for row in rows if row['sla_compliant'] is True)
        breached = sum(1 for row in rows if row['sla_compliant'] is False)
        return [
            {'Metric': 'Compliant', 'Value': compliant},
            {'Metric': 'Breached', 'Value': breached},
            {'Metric': 'Open overdue', 'Value': sum(1 for row in rows if row['is_overdue'])},
        ]

    def _scope_filters(self, filters: AnalyticsFilters, current_user) -> AnalyticsFilters:
        """
        Apply role-based reporting constraints to user-supplied filters.

        Purpose:
        Prevents support agents from viewing other agents' analytics.

        Parameters:
        filters: Requested report filters.
        current_user: Authenticated user.

        Returns:
        AnalyticsFilters with agent_id forced for support agents.

        Business Logic:
        Admin and operations users keep their requested agent filter; support
        agents are always scoped to their own user ID.
        """
        if is_support_agent(current_user):
            return AnalyticsFilters(
                start_date=filters.start_date,
                end_date=filters.end_date,
                agent_id=current_user.id,
                category_id=filters.category_id,
                status=filters.status,
            )
        return filters

    def _first_response_after(self, replies: list[Message], first_inbound_at: datetime | None) -> Message | None:
        """
        Find the first agent reply after the first buyer message.

        Purpose:
        Supports first-response and SLA calculations.

        Parameters:
        replies: Candidate outbound agent messages.
        first_inbound_at: Timestamp of the first buyer message.

        Returns:
        First qualifying Message, or None when no response exists.

        Business Logic:
        Replies before the first inbound buyer message are ignored because they
        cannot satisfy the buyer-response SLA.
        """
        if not first_inbound_at:
            return replies[0] if replies else None
        return next((message for message in replies if message.sent_at >= first_inbound_at), None)

    def _message_direction(self, message: Message | None) -> str | None:
        """
        Convert the latest message into a business direction label.

        Purpose:
        Feeds the Last Message Direction indicator in the inbox and reports.

        Parameters:
        message: Latest message in a conversation.

        Returns:
        Buyer, Agent, System, or None.

        Business Logic:
        Inbound customer/provider messages are treated as Buyer; outbound
        agent messages are Agent; explicit system sender types are System.
        """
        if not message:
            return None
        if message.sender_type == MessageSenderType.SYSTEM:
            return 'System'
        if message.sender_type == MessageSenderType.AGENT or not message.is_inbound:
            return 'Agent'
        return 'Buyer'

    def _reply_actor_name(self, message: Message) -> str | None:
        """
        Resolve a reply actor label for historical reporting.

        Purpose:
        Populates Replied By in exports using available local message data.

        Parameters:
        message: Agent reply message.

        Returns:
        Sender identifier captured with the message, or None.

        Business Logic:
        Existing reply rows store the seller account username as sender
        identifier, so this method preserves available history without a
        schema migration.
        """
        return message.sender_identifier

    def _daily_trends(self, rows: list[dict], filters: AnalyticsFilters) -> list[dict]:
        """
        Build daily conversation volume metrics.

        Purpose:
        Drives daily trend analysis charts.

        Parameters:
        rows: Report records produced by _conversation_record.
        filters: Active filters used to determine the date window.

        Returns:
        List of metric dictionaries ordered by date label.

        Business Logic:
        Counts conversations by latest message date when available, otherwise
        by creation date.
        """
        grouped: dict[str, int] = {}
        for row in rows:
            value = row['last_message_at'] or row['created_at']
            if value:
                key = value.date().isoformat()
                grouped[key] = grouped.get(key, 0) + 1
        return [self._metric(key, grouped[key]) for key in sorted(grouped)]

    def _count_by(self, rows: list[dict], key: str, default: str) -> list[dict]:
        """
        Count rows by a named field.

        Purpose:
        Produces category, status, and assignment distributions.

        Parameters:
        rows: Report records.
        key: Dictionary key to group by.
        default: Bucket label for empty values.

        Returns:
        Sorted metric dictionaries.

        Business Logic:
        Empty labels are grouped into an explicit default bucket.
        """
        grouped = self._group_rows(rows, key, default)
        return [self._metric(label, len(items)) for label, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0]))]

    def _sum_by(self, rows: list[dict], key: str, value_key: str, default: str) -> list[dict]:
        """
        Sum a numeric row field by group label.

        Purpose:
        Produces agent productivity reply counts.

        Parameters:
        rows: Report records.
        key: Grouping field.
        value_key: Numeric field to sum.
        default: Bucket label for empty values.

        Returns:
        Sorted metric dictionaries.

        Business Logic:
        Missing numeric values are treated as zero.
        """
        totals: dict[str, int | float] = {}
        for row in rows:
            label = row.get(key) or default
            totals[label] = totals.get(label, 0) + (row.get(value_key) or 0)
        return [self._metric(label, value) for label, value in sorted(totals.items(), key=lambda item: (-item[1], item[0]))]

    def _excel_datetime(self, value: datetime | None) -> datetime | None:
        """
        Convert timezone-aware datetimes into Excel-compatible naive values.

        Purpose:
        Prevents openpyxl from raising timezone errors during workbook export.

        Parameters:
        value: Optional datetime from database records.

        Returns:
        Naive UTC datetime or None.

        Business Logic:
        Excel cannot store timezone metadata. Values are normalized to UTC and
        written without tzinfo so chronological ordering remains stable.
        """
        if not value:
            return None
        if value.tzinfo:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value

    def _group_rows(self, rows: list[dict], key: str, default: str) -> dict[str, list[dict]]:
        """
        Group report records by a selected key.

        Purpose:
        Shares grouping behavior across API metrics and Excel summaries.

        Parameters:
        rows: Report records.
        key: Dictionary key used for grouping.
        default: Bucket label for empty values.

        Returns:
        Mapping of label to matching report records.

        Business Logic:
        Preserves every conversation by assigning empty values to the default
        bucket.
        """
        grouped: dict[str, list[dict]] = {}
        for row in rows:
            grouped.setdefault(row.get(key) or default, []).append(row)
        return grouped

    def _style_header(self, sheet) -> None:
        """
        Apply consistent report header styling.

        Purpose:
        Improves readability of generated Excel worksheets.

        Parameters:
        sheet: openpyxl worksheet with a header row.

        Returns:
        None.

        Business Logic:
        Header styling is presentation-only and does not affect report values.
        """
        fill = PatternFill(fill_type='solid', fgColor='1F2937')
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = fill
        for column in sheet.columns:
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = min(max(len(str(cell.value or '')) for cell in column) + 2, 36)

    def _metric(self, label: str, value) -> dict:
        """
        Create a dashboard metric object.

        Purpose:
        Keeps metric response shape consistent.

        Parameters:
        label: Human-readable metric label.
        value: Metric value.

        Returns:
        Dictionary with label and value keys.

        Business Logic:
        Values are passed through unchanged so counts, percentages, and labels
        can share one API shape.
        """
        return {'label': label, 'value': value}

    def _percent(self, numerator: int, denominator: int) -> str:
        """
        Format a percentage for reporting.

        Purpose:
        Presents SLA compliance in a dashboard-friendly form.

        Parameters:
        numerator: Count of compliant records.
        denominator: Count of measurable records.

        Returns:
        Percentage string with one decimal place, or N/A.

        Business Logic:
        Avoids dividing by zero when no SLA-measurable conversations exist.
        """
        if not denominator:
            return 'N/A'
        return f'{(numerator / denominator) * 100:.1f}%'

    def _minutes_label(self, minutes: float | None) -> str:
        """
        Convert response minutes into a compact label.

        Purpose:
        Displays average response time consistently.

        Parameters:
        minutes: Average minutes or None.

        Returns:
        Human-readable duration label.

        Business Logic:
        Uses hours for longer durations and minutes for short durations.
        """
        if minutes is None:
            return 'N/A'
        if minutes >= 60:
            return f'{minutes / 60:.1f}h'
        return f'{minutes:.0f}m'

    def _average(self, values: list[float]) -> float | None:
        """
        Calculate a numeric average.

        Purpose:
        Supports average response-time metrics.

        Parameters:
        values: Numeric values to average.

        Returns:
        Average value or None when the list is empty.

        Business Logic:
        Empty populations return None so callers can render N/A.
        """
        return sum(values) / len(values) if values else None

    def _start_datetime(self, value: date) -> datetime:
        """
        Convert a report start date to an inclusive UTC datetime.

        Purpose:
        Normalizes date-only filters for database comparison.

        Parameters:
        value: Start date.

        Returns:
        Datetime at 00:00:00 UTC.

        Business Logic:
        Date filters are inclusive for business reporting periods.
        """
        return datetime.combine(value, time.min, tzinfo=UTC)

    def _end_datetime(self, value: date) -> datetime:
        """
        Convert a report end date to an inclusive UTC datetime.

        Purpose:
        Normalizes date-only filters for database comparison.

        Parameters:
        value: End date.

        Returns:
        Datetime at 23:59:59.999999 UTC.

        Business Logic:
        Includes all activity on the selected end date.
        """
        return datetime.combine(value, time.max, tzinfo=UTC)

    def _role_scope(self, current_user) -> str:
        """
        Return the dashboard role scope label.

        Purpose:
        Allows the frontend to adjust headings for admin, operations, or agent
        analytics.

        Parameters:
        current_user: Authenticated user.

        Returns:
        ADMIN, OPERATIONS, or AGENT.

        Business Logic:
        Support-agent scoping is distinct because those dashboards are personal.
        """
        role_name = str(current_user.role.name if current_user and current_user.role else '').upper()
        if 'ADMIN' in role_name:
            return 'ADMIN'
        if 'SUPPORT' in role_name or 'AGENT' in role_name:
            return 'AGENT'
        return 'OPERATIONS'
