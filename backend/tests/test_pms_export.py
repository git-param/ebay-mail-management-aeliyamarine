from openpyxl import load_workbook

from app.modules.pms.export import export_monthly_table, export_monthly_tables
from app.modules.pms.schema import PmsMonthlyMetricSchema, PmsMonthlyTableResponse, PmsMonthlyTableRow


def metric(key: str, value: float, weight: float):
    return PmsMonthlyMetricSchema(
        metric_key=key,
        metric_name_snapshot=key,
        weight_snapshot=weight,
        source_snapshot='MANUAL',
        is_auto_calculated_snapshot=False,
        final_value=value,
    )


def test_pms_export_matches_reference_structure_and_blank_doj():
    table = PmsMonthlyTableResponse(
        year=2026,
        month=4,
        total_active_weight=100,
        completed_count=1,
        pending_count=0,
        items=[
            PmsMonthlyTableRow(
                user_id='11111111-1111-1111-1111-111111111111',
                user_name='Sample User',
                user_email='sample@example.com',
                status='COMPLETED',
                final_score=71,
                maximum_score=100,
                remarks='Solid month',
                metrics=[
                    metric('target_achievement', 45, 65),
                    metric('productivity', 8, 10),
                    metric('quality', 7, 10),
                    metric('punctuality', 4, 5),
                    metric('attendance', 3, 5),
                    metric('competencies', 4, 5),
                ],
            )
        ],
    )

    workbook = load_workbook(export_monthly_table(table), data_only=False)
    sheet = workbook.active

    assert sheet['A1'].value == 'PMS Monthly Data 26-27 For L-1 & L-2'
    assert 'A1:P1' in [str(item) for item in sheet.merged_cells.ranges]
    assert 'B2:H2' in [str(item) for item in sheet.merged_cells.ranges]
    assert 'I2:K2' in [str(item) for item in sheet.merged_cells.ranges]
    assert 'L2:N2' in [str(item) for item in sheet.merged_cells.ranges]
    assert sheet['A1'].fill.fgColor.rgb == '0017365D'
    assert sheet['B2'].fill.fgColor.rgb == '00DCE6F1'
    assert sheet['G5'].fill.fgColor.rgb == '00F2DCDB'
    assert sheet['O5'].fill.fgColor.rgb == '00E2F0C2'
    assert sheet['F5'].value is None
    assert sheet['G5'].value is None
    assert sheet['D5'].value == 'Operations'
    assert sheet['N5'].value == 4
    assert sheet['O5'].value == '=SUM(I5:N5)'
    assert sheet['P5'].value == 'Solid month'


def test_pms_export_supports_multiple_months_with_continuous_serial_numbers():
    first = PmsMonthlyTableResponse(
        year=2026,
        month=4,
        total_active_weight=100,
        completed_count=1,
        pending_count=0,
        items=[
            PmsMonthlyTableRow(
                user_id='11111111-1111-1111-1111-111111111111',
                user_name='April User',
                final_score=71,
                maximum_score=100,
                metrics=[metric('target_achievement', 45, 65)],
            )
        ],
    )
    second = PmsMonthlyTableResponse(
        year=2026,
        month=5,
        total_active_weight=100,
        completed_count=1,
        pending_count=0,
        items=[
            PmsMonthlyTableRow(
                user_id='22222222-2222-2222-2222-222222222222',
                user_name='May User',
                final_score=72,
                maximum_score=100,
                metrics=[metric('target_achievement', 46, 65)],
            )
        ],
    )

    workbook = load_workbook(export_monthly_tables([first, second]), data_only=False)
    sheet = workbook.active

    assert sheet['A1'].value == "PMS Monthly Data Apr'26 to May'26 For L-1 & L-2"
    assert sheet['A5'].value == 1
    assert sheet['H5'].value == "Apr'26"
    assert sheet['A6'].value == 2
    assert sheet['H6'].value == "May'26"
